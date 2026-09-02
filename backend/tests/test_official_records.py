"""The official FSSAI record source: what may enter, in what order, and what may match.

Every test here defends one property of an authoritative copy. A government
register is only useful if our copy cannot silently go stale, cannot be rewritten
by an older download, and cannot attach a stranger's recall to somebody's shelf.
"""
from __future__ import annotations

import asyncio
import hashlib
import json
import os
import subprocess
import sys
import zipfile
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from app.domains.official_records import service
from app.domains.official_records.matching import match_recall, resolve_matches
from app.domains.official_records.models import OfficialRecord, OfficialRecordRevision, OfficialSourceFetch
from app.domains.official_records.source import (
    BATCH_PLACEHOLDERS,
    HEADERS,
    SOURCE_ADAPTER_VERSION,
    SOURCE_ERROR_CODES,
    ZIP_LOCAL_FILE_HEADER,
    SourceError,
    normalise_batch,
    normalise_identity_text,
    normalise_licence,
    parse_recall_xlsx,
)
from app.shared.database.sql import get_sessionmaker
from openpyxl import Workbook
from sqlalchemy import select, text

LICENCE = "10012345678901"
FIXTURE = Path(__file__).parent / "fixtures" / "fssai_food_recall_v1.xlsx"
BACKEND_ROOT = Path(__file__).resolve().parent.parent


def data_row(*, recall_id=901, status="Initiated", batch="B-123", start="01-08-2026",
             termination="NA", licence=LICENCE, brand="Synthetic Brand", product="Synthetic cereal"):
    return [1, recall_id, " Synthetic Foods ", brand, batch, product, "Synthetic reason",
            start, status, termination, licence, "Central License", "Initiated by Authority"]


def make_export(path, *, rows=None, headers=HEADERS, sheet="data", formula=False, **changes):
    """A workbook shaped like the public export, with one knob per failure mode."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    worksheet.append(list(headers))
    for row in (rows if rows is not None else [data_row(**changes)]):
        worksheet.append(row)
    if formula:
        worksheet["G2"] = "=1+1"
    workbook.save(path)
    return path


def pack(**changes):
    return {"fssai_licence": LICENCE, "batch_number": "B-123", "brand": "Synthetic Brand",
            "product_name": "Synthetic cereal", **changes}


def official(**changes):
    return {"licence": LICENCE, "batch_lot": "B-123", "brand_name": "Synthetic Brand",
            "product_name": "Synthetic cereal", **changes}


def failure(path) -> str:
    with pytest.raises(SourceError) as caught:
        parse_recall_xlsx(Path(path))
    assert caught.value.code in SOURCE_ERROR_CODES
    return caught.value.code


# ---------------------------------------------------------------------------
# The workbook contract
# ---------------------------------------------------------------------------

def test_zip_signature_is_four_real_bytes_not_an_escaped_literal():
    """Regression: ``b"PK\\\\x03\\\\x04"`` is ten characters of literal backslash.

    Spelling the signature that way compiles, reads correctly at a glance, and
    rejects every genuine XLSX ever produced. Only the byte values defend it.
    """
    assert bytes((0x50, 0x4B, 0x03, 0x04)) == ZIP_LOCAL_FILE_HEADER
    assert len(ZIP_LOCAL_FILE_HEADER) == 4
    assert b"\\" not in ZIP_LOCAL_FILE_HEADER
    assert FIXTURE.read_bytes().startswith(ZIP_LOCAL_FILE_HEADER)


def test_public_xlsx_contract_preserves_identifiers_dates_and_original_bytes():
    rows, digest = parse_recall_xlsx(FIXTURE)
    assert digest == hashlib.sha256(FIXTURE.read_bytes()).hexdigest()
    assert SOURCE_ADAPTER_VERSION.endswith("xlsx.v1")
    assert len(rows) == 2
    assert rows[0]["external_record_id"] == "901"
    assert rows[0]["fbo_name"] == "Synthetic Foods Private Limited"
    assert rows[0]["licence"] == LICENCE
    assert rows[0]["batch_lot"] == "B-123"
    assert rows[0]["recall_start_date"].isoformat() == "2026-08-01"
    # "NA" is the register saying the recall has not ended, not a date.
    assert rows[0]["recall_termination_date"] is None
    assert rows[1]["recall_termination_date"].isoformat() == "2026-08-03"


def test_openpyxl_written_workbook_parses_and_collapses_whitespace(tmp_path):
    rows, digest = parse_recall_xlsx(make_export(tmp_path / "synthetic.xlsx"))
    assert digest == hashlib.sha256((tmp_path / "synthetic.xlsx").read_bytes()).hexdigest()
    assert [row["fbo_name"] for row in rows] == ["Synthetic Foods"]
    assert rows[0]["external_record_id"] == "901"


def test_header_and_worksheet_contract_is_exact(tmp_path):
    assert failure(make_export(tmp_path / "sheet.xlsx", sheet="Sheet1")) == "unsupported_official_export"
    renamed = list(HEADERS[:-1]) + ["Nature Of Recall"]
    assert failure(make_export(tmp_path / "renamed.xlsx", headers=renamed)) == "unexpected_official_export_schema"
    assert failure(make_export(tmp_path / "short.xlsx", headers=HEADERS[:-1],
                               rows=[data_row()[:-1]])) == "unexpected_official_export_schema"
    extra = [*HEADERS, "Inspector Notes"]
    assert failure(make_export(tmp_path / "extra.xlsx", headers=extra,
                               rows=[[*data_row(), "n/a"]])) == "unexpected_official_export_schema"


def test_second_worksheet_is_refused(tmp_path):
    path = tmp_path / "two-sheets.xlsx"
    make_export(path)
    from openpyxl import load_workbook
    workbook = load_workbook(path)
    workbook.create_sheet("notes")
    workbook.save(path)
    assert failure(path) == "unsupported_official_export"


def test_header_only_export_is_refused_rather_than_recorded_as_a_successful_check(tmp_path):
    """An empty success would advance freshness and make a stale copy look current."""
    assert failure(make_export(tmp_path / "empty.xlsx", rows=[])) == "empty_official_export"


def test_duplicate_recall_id_fails_the_whole_artifact(tmp_path):
    rows = [data_row(recall_id=901, status="Initiated"), data_row(recall_id=901, status="Completed")]
    assert failure(make_export(tmp_path / "duplicate.xlsx", rows=rows)) == "duplicate_official_record_id"


def test_missing_recall_id_is_refused(tmp_path):
    assert failure(make_export(tmp_path / "no-id.xlsx", recall_id=None)) == "missing_official_record_id"


def test_malformed_dates_fail_on_either_date_column(tmp_path):
    assert failure(make_export(tmp_path / "start.xlsx", start="2026-08-01")) == "malformed_official_date"
    assert failure(make_export(tmp_path / "end.xlsx", termination="2026/08/03")) == "malformed_official_date"


def test_formula_cell_is_refused(tmp_path):
    assert failure(make_export(tmp_path / "formula.xlsx", formula=True)) == "invalid_official_export"


def test_numeric_identifier_cells_are_refused_before_str_coercion(tmp_path):
    """A numeric licence cell would reach the database as ``10012345678901.0``."""
    assert failure(make_export(tmp_path / "licence.xlsx", licence=10012345678901)) == "malformed_official_identifier"
    # A numeric batch loses its leading zeros the same way.
    assert failure(make_export(tmp_path / "batch.xlsx", batch=789)) == "malformed_official_identifier"


def test_identifier_type_lock_covers_every_way_a_cell_can_stop_being_text(tmp_path):
    """Booleans and fractions are not identifiers; a blank cell is simply absent."""
    assert failure(make_export(tmp_path / "bool.xlsx", licence=True)) == "malformed_official_identifier"
    assert failure(make_export(tmp_path / "float-id.xlsx", recall_id=901.5)) == "malformed_official_identifier"
    # An integral numeric Recall Id is the real observed source type.
    rows, _ = parse_recall_xlsx(make_export(tmp_path / "float-ok.xlsx", recall_id=901.0))
    assert rows[0]["external_record_id"] == "901"
    blank, _ = parse_recall_xlsx(make_export(tmp_path / "blank.xlsx", licence="   "))
    assert blank[0]["licence"] is None


def test_trailing_blank_rows_are_not_recall_records(tmp_path):
    """Spreadsheets carry empty trailing rows; they are not an official record."""
    rows, _ = parse_recall_xlsx(make_export(
        tmp_path / "trailing.xlsx", rows=[data_row(), [None] * len(HEADERS), [None] * len(HEADERS)],
    ))
    assert [row["external_record_id"] for row in rows] == ["901"]


def test_non_xlsx_bytes_and_containers_are_refused(tmp_path):
    csv = tmp_path / "renamed.xlsx"
    csv.write_text(",".join(HEADERS) + "\n1,901,Synthetic Foods\n", encoding="utf-8")
    assert failure(csv) == "unsupported_official_export"

    corrupt = tmp_path / "corrupt.xlsx"
    good = make_export(tmp_path / "good.xlsx").read_bytes()
    corrupt.write_bytes(good[:len(good) // 2])
    assert failure(corrupt) == "unsupported_official_export"

    truncated_directory = tmp_path / "truncated.xlsx"
    truncated_directory.write_bytes(ZIP_LOCAL_FILE_HEADER + b"not a real archive")
    assert failure(truncated_directory) == "unsupported_official_export"

    wrong_suffix = tmp_path / "export.xls"
    wrong_suffix.write_bytes(good)
    assert failure(wrong_suffix) == "unsupported_official_export"


def test_macro_package_renamed_to_xlsx_is_refused(tmp_path):
    """openpyxl only fills ``vba_archive`` when loaded with keep_vba, so the
    package itself is inspected. A .xlsm renamed .xlsx is still macro-enabled."""
    source = make_export(tmp_path / "plain.xlsx")
    macro = tmp_path / "macro.xlsx"
    with zipfile.ZipFile(source) as original, zipfile.ZipFile(macro, "w") as rebuilt:
        for item in original.infolist():
            rebuilt.writestr(item, original.read(item.filename))
        rebuilt.writestr("xl/vbaProject.bin", b"\xd0\xcf\x11\xe0synthetic macro part")
    assert failure(macro) == "macro_enabled_official_export"


def test_failed_parse_keeps_the_original_file_digest_where_readable(tmp_path):
    path = make_export(tmp_path / "bad-date.xlsx", start="2026-08-01")
    with pytest.raises(SourceError) as caught:
        parse_recall_xlsx(path)
    assert caught.value.source_file_sha256 == hashlib.sha256(path.read_bytes()).hexdigest()


# ---------------------------------------------------------------------------
# Placeholder batches, and the identity text that is not a batch
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("placeholder", [
    "NA", "N/A", "nil", "none", "Not Applicable", "not available", "other", "others",
    "-", ".", "NO", "LOOSE", "loose sample", "sold as loose", "0", "00", "000", "  NA  ",
])
def test_source_placeholder_batches_are_not_matchable(placeholder):
    assert normalise_batch(placeholder) is None
    assert match_recall(pack(batch_number=placeholder), official(batch_lot=placeholder)) == "not_matched"


@pytest.mark.parametrize("batch", ["C", "1", "B-123", "0789/55", "L0"])
def test_real_short_batches_stay_meaningful(batch):
    assert normalise_batch(batch) == batch.casefold()
    assert match_recall(pack(batch_number=batch), official(batch_lot=batch)) == "matched"


@pytest.mark.parametrize("licence", ["10012345678901", "  10012345678901  ", "1001 2345 6789 01"])
def test_a_licence_is_its_fourteen_printed_digits(licence):
    """Labels print the number grouped; the digits themselves are the identifier."""
    assert normalise_licence(licence) == LICENCE
    assert match_recall(pack(fssai_licence=licence), official()) == "matched"


@pytest.mark.parametrize("malformed", [
    "FSSAI 10012345678901", "10012345678901X", "X10012345678901", "10012-345678901",
    "1001234567890", "100123456789012", "00000000000000", "1001234567890a",
])
def test_a_licence_is_never_salvaged_by_deleting_characters(malformed):
    """Stripping non-digits would invent an exactness the source never stated."""
    assert normalise_licence(malformed) is None
    assert match_recall(pack(fssai_licence=malformed), official()) == "not_matched"
    # Two equally malformed values are not evidence of the same manufacturer.
    assert match_recall(pack(fssai_licence=malformed), official(licence=malformed)) == "not_matched"


def test_separators_are_not_stripped_so_similar_lots_stay_distinct():
    assert normalise_batch("B-123") != normalise_batch("B 123")
    assert match_recall(pack(batch_number="B-123"), official(batch_lot="B 123")) == "identity_mismatch"


def test_identity_normalisation_does_not_inherit_the_batch_placeholder_vocabulary():
    """A brand may legitimately be called ``Other``; a lot number may not."""
    for value in ("Other", "None", "NO", "0", "00"):
        assert normalise_batch(value) is None
        assert normalise_identity_text(value) == value.casefold()
    assert normalise_identity_text("  Nature's   Own  ") == "nature's own"
    assert normalise_identity_text(None) is None
    assert normalise_identity_text("   ") is None
    assert "other" in BATCH_PLACEHOLDERS


def test_exact_licence_and_batch_decide_eligibility_and_identity_only_guards_conflict():
    assert match_recall(pack(), official()) == "matched"
    # A brand named with the batch vocabulary is still a brand.
    assert match_recall(pack(brand="Other"), official(brand_name="other")) == "matched"
    # Missing identity on either side is missing information, not disagreement.
    assert match_recall(pack(brand=None, product_name=None), official()) == "matched"
    assert match_recall(pack(), official(brand_name=None, product_name=None)) == "matched"
    # A real conflict still blocks the match.
    assert match_recall(pack(brand="Different Brand"), official()) == "identity_mismatch"
    assert match_recall(pack(product_name="Different product"), official()) == "identity_mismatch"
    # Identity alone never establishes a match without licence and batch.
    assert match_recall(pack(fssai_licence=None), official()) == "not_matched"
    assert match_recall(pack(batch_number=None), official()) == "not_matched"
    assert match_recall(pack(), official(licence=None)) == "not_matched"
    assert match_recall(pack(), official(batch_lot=None)) == "not_matched"
    # An invalid licence shape is not an identifier.
    assert match_recall(pack(fssai_licence="123"), official(licence="123")) == "not_matched"
    assert match_recall(pack(fssai_licence="0" * 14), official(licence="0" * 14)) == "not_matched"


# ---------------------------------------------------------------------------
# Several rows can share one licence and lot. Resolving them is a set problem.
# ---------------------------------------------------------------------------

R1 = {"recall_id": "R1", "licence": LICENCE, "batch_lot": "B-123", "brand_name": "Alpha", "product_name": "Oats"}
R2 = {"recall_id": "R2", "licence": LICENCE, "batch_lot": "B-123", "brand_name": "Beta", "product_name": "Juice"}
R3 = {"recall_id": "R3", "licence": LICENCE, "batch_lot": "B-123", "brand_name": "Alpha", "product_name": "Oats"}


def resolved(pack_facts, records):
    return [record["recall_id"] for record in resolve_matches(pack_facts, records)]


def test_a_pack_without_identity_does_not_inherit_every_recall_on_its_licence():
    """Case A. One licence and lot can name several products; a pack is one of them.

    Handing all of them to the customer would present a stranger's recall as
    theirs, and picking one would be a guess dressed as a fact.
    """
    assert resolved(pack(brand=None, product_name=None), [R1, R2]) == []


def test_exact_pack_identity_selects_exactly_its_own_record():
    """Cases B and C."""
    assert resolved(pack(brand="Alpha", product_name="Oats"), [R1, R2]) == ["R1"]
    assert resolved(pack(brand="Beta", product_name="Juice"), [R1, R2]) == ["R2"]
    # Case D: identity that agrees with neither resolves to nothing.
    assert resolved(pack(brand="Gamma", product_name="Tea"), [R1, R2]) == []


def test_several_records_may_be_returned_only_for_one_corroborated_identity():
    """Case 3. Two filings against the same product are both that product's."""
    assert resolved(pack(brand="Alpha", product_name="Oats"), [R1, R3]) == ["R1", "R3"]
    # Case 4: a pack naming only its brand cannot choose between two products.
    assert resolved(pack(brand="Alpha", product_name=None),
                    [R1, {**R2, "brand_name": "Alpha"}]) == []
    # The same two rows without pack identity are still unresolved, not "both".
    assert resolved(pack(brand=None, product_name=None), [R1, R3]) == []
    # A single candidate never needs corroboration; nothing else could be meant.
    assert resolved(pack(brand=None, product_name=None), [R1]) == ["R1"]


def test_a_candidate_the_register_leaves_unidentified_keeps_the_set_unresolved():
    """Case 1. Missing official identity is uncertainty, not evidence against a row.

    The register filed a second recall on this licence and lot and named no
    brand or product for it. That row cannot be confirmed as this pack's — and
    it cannot be excluded either. Publishing the row we *could* confirm, while
    that one sits beside it unresolved, would present a guess as a fact.
    """
    blank = {"recall_id": "R0", "licence": LICENCE, "batch_lot": "B-123",
             "brand_name": None, "product_name": None}
    assert resolved(pack(brand="Alpha", product_name="Oats"), [R1, blank]) == []
    # Alone, that same row is still the only thing the licence and lot can mean.
    assert resolved(pack(brand="Alpha", product_name="Oats"), [blank]) == ["R0"]


def test_a_partly_identified_candidate_is_not_completed_by_inference():
    """Case 2. A row naming only the product is not thereby the same product identity.

    Filling in the missing brand from the other candidate would be inventing the
    fact that decides whether these two rows are one product or two.
    """
    partial = {"recall_id": "R0", "licence": LICENCE, "batch_lot": "B-123",
               "brand_name": None, "product_name": "Oats"}
    assert resolved(pack(brand="Alpha", product_name="Oats"), [R1, partial]) == []


def test_ruled_out_and_unidentified_candidates_are_not_the_same_thing():
    """Case 5 against Case 1. This distinction decides what a customer is shown."""
    conflicting = {"recall_id": "R0", "licence": LICENCE, "batch_lot": "B-123",
                   "brand_name": "Beta", "product_name": "Juice"}
    unidentified = {**conflicting, "brand_name": None, "product_name": None}
    # A row stating a different brand has been ruled out; the set resolves.
    assert resolved(pack(brand="Alpha", product_name="Oats"), [R1, conflicting]) == ["R1"]
    # A row stating no brand has not been ruled out; the set stays unresolved.
    assert resolved(pack(brand="Alpha", product_name="Oats"), [R1, unidentified]) == []


def test_resolution_never_falls_back_to_first_newest_or_all():
    ambiguous = [R1, R2, R3]
    assert resolved(pack(brand=None, product_name=None), ambiguous) == []
    # Corroborating only one side is enough when it names one identity group.
    assert resolved(pack(brand="Alpha", product_name=None), ambiguous) == ["R1", "R3"]
    # Corroborating a side that spans two identities stays unresolved.
    spanning = [{**R1, "recall_id": "R4", "product_name": "Muesli"}, R1]
    assert resolved(pack(brand="Alpha", product_name=None), spanning) == []
    # An unusable licence resolves to nothing however many rows agree.
    assert resolved(pack(fssai_licence="FSSAI 10012345678901"), ambiguous) == []


# ---------------------------------------------------------------------------
# Source order: an authoritative copy only ever moves forward
# ---------------------------------------------------------------------------

T0 = datetime(2026, 7, 20, tzinfo=UTC)
T1 = datetime(2026, 8, 1, tzinfo=UTC)
T2 = datetime(2026, 8, 4, tzinfo=UTC)
T3 = datetime(2026, 8, 9, tzinfo=UTC)


async def _ingest(path, checked_at):
    factory = get_sessionmaker()
    async with factory() as session:
        fetch, counts = await service.ingest_recall_xlsx(session, path, source_checked_at=checked_at)
        await session.commit()
        return fetch, counts


async def _reject(path, checked_at) -> str:
    factory = get_sessionmaker()
    async with factory() as session:
        with pytest.raises(SourceError) as caught:
            await service.ingest_recall_xlsx(session, path, source_checked_at=checked_at)
        await session.rollback()
        return caught.value.code


async def _state():
    factory = get_sessionmaker()
    async with factory() as session:
        records = (await session.execute(select(OfficialRecord))).scalars().all()
        revisions = (await session.execute(
            select(OfficialRecordRevision).order_by(OfficialRecordRevision.revision_number)
        )).scalars().all()
        freshness = (await service.official_records_envelope(session, None))["last_successful_check_at"]
        return records, revisions, freshness


@pytest.mark.anyio
async def test_unchanged_reobservation_advances_observation_without_inventing_a_revision(tmp_path, db_clean):
    """Seeing the same record again is not the register changing its mind."""
    first, second = tmp_path / "t1.xlsx", tmp_path / "t2.xlsx"
    make_export(first)
    make_export(second)
    fetch_one, created = await _ingest(first, T1)
    fetch_two, again = await _ingest(second, T2)
    records, revisions, freshness = await _state()
    assert created["records_created"] == 1
    assert again == {"records_created": 0, "records_revised": 0, "records_unchanged": 1}
    assert len(records) == 1 and records[0].latest_revision == 1
    assert len(revisions) == 1 and revisions[0].source_fetch_id == fetch_one.id
    assert records[0].last_seen_at == T2
    assert records[0].last_seen_fetch_id == fetch_two.id
    assert records[0].first_seen_at == T1
    assert freshness == T2.isoformat()


@pytest.mark.anyio
async def test_source_order_governs_revisions_freshness_and_regression(tmp_path, db_clean):
    initial = make_export(tmp_path / "t1.xlsx")
    changed = make_export(tmp_path / "t2.xlsx", status="Completed")
    unchanged = make_export(tmp_path / "t3.xlsx", status="Completed")
    older = make_export(tmp_path / "t0.xlsx", status="Initiated")

    fetch_one, created = await _ingest(initial, T1)
    fetch_two, revised = await _ingest(changed, T2)
    fetch_three, seen_again = await _ingest(unchanged, T3)
    assert created["records_created"] == 1
    assert revised["records_revised"] == 1
    assert seen_again["records_unchanged"] == 1

    records, revisions, freshness = await _state()
    assert records[0].latest_revision == 2
    assert [revision.source_fetch_id for revision in revisions] == [fetch_one.id, fetch_two.id]
    assert records[0].recall_status == "Completed"
    assert records[0].last_seen_at == T3 and records[0].last_seen_fetch_id == fetch_three.id
    assert freshness == T3.isoformat()

    # An older download must not rewrite the register with content it has superseded.
    assert await _reject(older, T0) == "out_of_order_source_check"
    after_records, after_revisions, after_freshness = await _state()
    assert after_records[0].recall_status == "Completed"
    assert after_records[0].latest_revision == 2
    assert after_records[0].last_seen_at == T3
    assert after_records[0].last_seen_fetch_id == fetch_three.id
    assert len(after_revisions) == 2
    assert after_freshness == T3.isoformat()


@pytest.mark.anyio
async def test_a_later_status_is_never_overwritten_by_an_earlier_source_check(tmp_path, db_clean):
    """T2 says Completed. A T1 file saying Initiated may not undo that."""
    completed = make_export(tmp_path / "t2.xlsx", status="Completed")
    initiated = make_export(tmp_path / "t1.xlsx", status="Initiated")
    fetch_two, _ = await _ingest(completed, T2)

    assert await _reject(initiated, T1) == "out_of_order_source_check"

    records, revisions, freshness = await _state()
    assert records[0].recall_status == "Completed"
    assert records[0].latest_revision == 1
    assert records[0].last_seen_at == T2
    assert records[0].last_seen_fetch_id == fetch_two.id
    assert len(revisions) == 1
    assert freshness == T2.isoformat()


@pytest.mark.anyio
async def test_equal_source_check_times_are_refused_deterministically(tmp_path, db_clean):
    """V1 picks no winner between two artifacts claiming the same instant.

    Identical bytes are an idempotent duplicate and change nothing; different
    bytes are a genuine conflict. Both are refused, and both leave the register
    exactly as the accepted check left it.
    """
    accepted = make_export(tmp_path / "t1.xlsx")
    fetch_one, _ = await _ingest(accepted, T1)
    before = await _state()

    identical = tmp_path / "identical.xlsx"
    identical.write_bytes(accepted.read_bytes())
    assert await _reject(identical, T1) == "duplicate_source_check"

    differing = make_export(tmp_path / "differing.xlsx", status="Completed")
    assert await _reject(differing, T1) == "conflicting_source_check"

    records, revisions, freshness = await _state()
    assert records[0].recall_status == before[0][0].recall_status == "Initiated"
    assert records[0].latest_revision == 1 and len(revisions) == 1
    assert records[0].last_seen_fetch_id == fetch_one.id
    assert freshness == T1.isoformat()


@pytest.mark.anyio
async def test_failed_source_is_recorded_without_advancing_successful_freshness(tmp_path, db_clean):
    good = make_export(tmp_path / "t1.xlsx")
    await _ingest(good, T1)
    malformed = make_export(tmp_path / "t2.xlsx", start="2026-08-04")

    factory = get_sessionmaker()
    async with factory() as session:
        with pytest.raises(SourceError) as caught:
            await service.ingest_recall_xlsx(session, malformed, source_checked_at=T2)
        await session.rollback()
        await service.record_source_error(session, caught.value, source_checked_at=T2, path=malformed)
        await session.commit()

    async with factory() as session:
        fetches = (await session.execute(
            select(OfficialSourceFetch).order_by(OfficialSourceFetch.source_checked_at)
        )).scalars().all()
    assert [row.status for row in fetches] == ["succeeded", "failed"]
    failed = fetches[1]
    assert failed.error_code == "malformed_official_date"
    # A rejected artifact stays auditable: which file, in what format, and its bytes.
    assert failed.original_filename == "t2.xlsx"
    assert failed.source_format == "xlsx"
    assert failed.source_file_sha256 == hashlib.sha256(malformed.read_bytes()).hexdigest()
    assert failed.source_checked_at == T2

    records, revisions, freshness = await _state()
    assert records[0].recall_status == "Initiated"
    assert len(revisions) == 1
    assert freshness == T1.isoformat()


@pytest.mark.anyio
async def test_error_codes_written_to_the_ledger_stay_inside_the_closed_vocabulary(tmp_path, db_clean):
    factory = get_sessionmaker()
    for name, kwargs in (("bad-date.xlsx", {"start": "x"}), ("bad-licence.xlsx", {"licence": 123}),
                         ("empty.xlsx", {"rows": []})):
        path = make_export(tmp_path / name, **kwargs)
        async with factory() as session:
            with pytest.raises(SourceError) as caught:
                await service.ingest_recall_xlsx(session, path, source_checked_at=T1)
            await session.rollback()
            await service.record_source_error(session, caught.value, source_checked_at=T1, path=path)
            await session.commit()
    async with factory() as session:
        codes = (await session.execute(select(OfficialSourceFetch.error_code))).scalars().all()
    assert set(codes) <= SOURCE_ERROR_CODES
    # No openpyxl, zipfile or XML parser text ever reaches the ledger.
    assert all(code and len(code) <= 64 and " " not in code for code in codes)


# ---------------------------------------------------------------------------
# Concurrency: the chronology guard must hold against a real second session
# ---------------------------------------------------------------------------

async def _ingest_concurrently(path, checked_at, start: asyncio.Event | None = None) -> str:
    """One import in its own session, released together with the others.

    The connection is opened and its transaction started before the barrier. An
    unwarmed session spends its first milliseconds on TCP and authentication,
    which is long enough for the other import to finish — the race would then
    never actually happen and the test would prove nothing.
    """
    factory = get_sessionmaker()
    async with factory() as session:
        await session.execute(text("SELECT 1"))
        if start is not None:
            await start.wait()
        try:
            await service.ingest_recall_xlsx(session, path, source_checked_at=checked_at)
            await session.commit()
            return "succeeded"
        except SourceError as exc:
            await session.rollback()
            return exc.code


async def _race(*attempts) -> list[str]:
    start = asyncio.Event()
    tasks = [asyncio.create_task(_ingest_concurrently(path, when, start)) for path, when in attempts]
    start.set()
    return await asyncio.gather(*tasks)


async def _fetches():
    factory = get_sessionmaker()
    async with factory() as session:
        return (await session.execute(
            select(OfficialSourceFetch).order_by(OfficialSourceFetch.created_at)
        )).scalars().all()


@pytest.mark.anyio
async def test_a_second_import_waits_on_the_database_lock_instead_of_reading_stale_state(tmp_path, db_clean):
    """The serialization is PostgreSQL's, so it holds across sessions and processes.

    While one transaction holds the source lock, a second import cannot even
    reach the "latest successful check" its chronology guard depends on. When it
    finally does, it reads the truth the first import committed.
    """
    await _ingest(make_export(tmp_path / "t1.xlsx"), T1)
    later = make_export(tmp_path / "t2.xlsx")
    latest = make_export(tmp_path / "t3.xlsx")

    factory = get_sessionmaker()
    async with factory() as holder:
        await service.lock_official_source(holder)
        blocked = asyncio.create_task(_ingest_concurrently(later, T2))
        # Long enough for an unserialized import to have finished several times.
        await asyncio.sleep(0.25)
        assert not blocked.done(), "a second import proceeded while the source lock was held"
        await service.ingest_recall_xlsx(holder, latest, source_checked_at=T3)
        await holder.commit()

    # Released, the waiting import now sees T3 and refuses to go backwards.
    assert await blocked == "out_of_order_source_check"
    records, revisions, freshness = await _state()
    assert freshness == T3.isoformat()
    assert records[0].last_seen_at == T3
    assert len(revisions) == 1


@pytest.mark.anyio
@pytest.mark.parametrize("newest_first", [False, True])
async def test_concurrent_imports_cannot_leave_older_content_under_newer_freshness(
    tmp_path, db_clean, newest_first,
):
    """Without database serialization both sessions read the same "latest check",
    both pass the chronology guard, and both mutate — so the guard's promise
    would be false exactly when it matters.

    Both arrival orders are exercised. The newest-first order is the one that
    breaks unserialized: T3 commits, then T2 commits over it, leaving the
    register holding T2's observation while freshness already reports T3.
    """
    await _ingest(make_export(tmp_path / "t1.xlsx"), T1)
    later = make_export(tmp_path / "t2.xlsx")
    latest = make_export(tmp_path / "t3.xlsx")
    attempts = [(latest, T3), (later, T2)] if newest_first else [(later, T2), (latest, T3)]

    outcomes = sorted(await _race(*attempts))
    # Either order is legitimate; T2 arriving after T3 is simply out of order.
    assert outcomes in (["out_of_order_source_check", "succeeded"], ["succeeded", "succeeded"]), outcomes

    records, revisions, freshness = await _state()
    assert freshness == T3.isoformat()
    assert records[0].last_seen_at == T3
    # Unchanged content never earns a second revision, however the race ran.
    assert records[0].latest_revision == 1
    assert len(revisions) == 1
    successful = [row for row in await _fetches() if row.status == "succeeded"]
    assert max(row.source_checked_at for row in successful) == T3


@pytest.mark.anyio
async def test_a_first_import_race_on_identical_bytes_admits_exactly_one(tmp_path, db_clean):
    """Two operators importing the same download at the same stated instant."""
    original = make_export(tmp_path / "foscos.xlsx")
    copy = tmp_path / "foscos-copy.xlsx"
    copy.write_bytes(original.read_bytes())

    assert sorted(await _race((original, T1), (copy, T1))) == ["duplicate_source_check", "succeeded"]

    records, revisions, _ = await _state()
    assert len(records) == 1 and len(revisions) == 1
    fetches = await _fetches()
    assert [row.status for row in fetches].count("succeeded") == 1


@pytest.mark.anyio
async def test_a_first_import_race_on_differing_bytes_picks_no_winner(tmp_path, db_clean):
    """Two artifacts claiming the same instant disagree; only one can be accepted."""
    first = make_export(tmp_path / "a.xlsx", status="Initiated")
    second = make_export(tmp_path / "b.xlsx", status="Completed")

    assert sorted(await _race((first, T1), (second, T1))) == ["conflicting_source_check", "succeeded"]

    records, revisions, freshness = await _state()
    assert len(records) == 1 and len(revisions) == 1
    assert freshness == T1.isoformat()
    fetches = await _fetches()
    assert [row.status for row in fetches].count("succeeded") == 1
    rejected = [row for row in fetches if row.status == "failed"]
    assert all(row.error_code in SOURCE_ERROR_CODES for row in rejected)


# ---------------------------------------------------------------------------
# A record we keep after it leaves the export
# ---------------------------------------------------------------------------

@pytest.mark.anyio
async def test_a_record_absent_from_the_latest_export_keeps_its_own_observation_date(tmp_path, db_clean):
    """Absence from one download proves nothing, so the record stays — dated honestly.

    The screen must be able to say "last observed on 1 August" while also saying
    "we last checked on 4 August". Collapsing those two into one date would claim
    an observation that never happened.
    """
    first = make_export(tmp_path / "t1.xlsx", recall_id=901)
    second = make_export(tmp_path / "t2.xlsx", recall_id=902)
    fetch_one, _ = await _ingest(first, T1)
    fetch_two, second_counts = await _ingest(second, T2)
    assert second_counts["records_created"] == 1

    factory = get_sessionmaker()
    async with factory() as session:
        records = {row.external_record_id: row for row in (
            await session.execute(select(OfficialRecord))
        ).scalars().all()}
        revisions = (await session.execute(select(OfficialRecordRevision))).scalars().all()
        envelope = await service.official_records_envelope(session, pack())
    # The 1 August record is retained, with its history and its own last-seen state.
    assert set(records) == {"901", "902"}
    assert records["901"].last_seen_at == T1
    assert records["901"].last_seen_fetch_id == fetch_one.id
    assert records["902"].last_seen_fetch_id == fetch_two.id
    assert len(revisions) == 2

    assert envelope["last_successful_check_at"] == T2.isoformat()
    matched = {row["recall_id"]: row for row in envelope["records"]}
    assert set(matched) == {"901", "902"}
    assert matched["901"]["source_last_seen_at"] == T1.isoformat()
    assert matched["901"]["seen_in_latest_successful_check"] is False
    # A record that was in the latest export says so, and dates itself to it.
    assert matched["902"]["source_last_seen_at"] == T2.isoformat()
    assert matched["902"]["seen_in_latest_successful_check"] is True
    # Nothing in the payload characterises the absence.
    forbidden = ("withdrawn", "removed", "no longer", "resolved", "cleared", "safe")
    assert not [word for word in forbidden if word in str(envelope).casefold()]


# ---------------------------------------------------------------------------
# The operator command
# ---------------------------------------------------------------------------

def run_command(*args) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, "-m", "app.commands.ingest_fssai_recalls", *args],
        cwd=BACKEND_ROOT, env={**os.environ}, capture_output=True, text=True, timeout=180,
    )


@pytest.mark.anyio
async def test_operator_command_reports_source_provenance_on_success(tmp_path, db_clean):
    path = make_export(tmp_path / "foscos.xlsx")
    completed = run_command(str(path), "--source-checked-at", "2026-08-01T10:00:00+05:30")
    assert completed.returncode == 0, completed.stderr
    report = json.loads(completed.stdout)
    assert set(report) == {
        "source_fetch_id", "records_in_export", "records_created", "records_revised",
        "records_unchanged", "source_checked_at", "source_file_sha256",
    }
    assert report["records_in_export"] == 1
    assert report["records_created"] == 1
    assert report["source_checked_at"] == "2026-08-01T10:00:00+05:30"
    assert report["source_file_sha256"] == hashlib.sha256(path.read_bytes()).hexdigest()


@pytest.mark.anyio
async def test_operator_command_fails_loudly_and_leaves_prior_good_data_alone(tmp_path, db_clean):
    good = make_export(tmp_path / "good.xlsx")
    assert run_command(str(good), "--source-checked-at", "2026-08-01T00:00:00+00:00").returncode == 0
    before = await _state()

    duplicate_ids = make_export(
        tmp_path / "duplicate.xlsx",
        rows=[data_row(recall_id=901), data_row(recall_id=901, status="Completed")],
    )
    failed = run_command(str(duplicate_ids), "--source-checked-at", "2026-08-04T00:00:00+00:00")
    assert failed.returncode == 1
    assert json.loads(failed.stderr) == {"error": "duplicate_official_record_id"}

    factory = get_sessionmaker()
    async with factory() as session:
        ledger = (await session.execute(
            select(OfficialSourceFetch).where(OfficialSourceFetch.status == "failed")
        )).scalars().all()
    assert len(ledger) == 1
    assert ledger[0].error_code == "duplicate_official_record_id"
    assert ledger[0].original_filename == "duplicate.xlsx"
    assert ledger[0].source_file_sha256 == hashlib.sha256(duplicate_ids.read_bytes()).hexdigest()

    after = await _state()
    assert [row.recall_status for row in after[0]] == [row.recall_status for row in before[0]]
    assert len(after[1]) == len(before[1])
    # A failed check never claims the copy is fresher than the last good one.
    assert after[2] == before[2] == "2026-08-01T00:00:00+00:00"


def test_operator_command_refuses_an_unstated_naive_or_future_source_time(tmp_path):
    path = make_export(tmp_path / "foscos.xlsx")
    # The download time is never inferred from the file or its name.
    assert run_command(str(path)).returncode == 2
    assert run_command(str(path), "--source-checked-at", "2026-08-01T10:00:00").returncode == 2
    ahead = (datetime.now(UTC) + timedelta(days=2)).isoformat()
    assert run_command(str(path), "--source-checked-at", ahead).returncode == 2
    assert run_command(str(path), "--source-checked-at", "the first of August").returncode == 2
