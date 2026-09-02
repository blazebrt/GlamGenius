from __future__ import annotations

import hashlib
import json
import unicodedata
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.domains.product.fssai import LICENCE_LENGTH, is_valid_licence

AUTHORITY_FSSAI_FOSCOS = "fssai_foscos"
RECORD_TYPE_FOOD_RECALL = "food_recall"
SOURCE_ADAPTER_VERSION = "fssai-foscos-food-recall.xlsx.v1"
SOURCE_URL = "https://foscos.fssai.gov.in/food-recall"
SOURCE_FORMAT = "xlsx"
MAX_SOURCE_BYTES = 10 * 1024 * 1024
MAX_SOURCE_ROWS = 10_000
SHEET_NAME = "data"

#: The four-byte local file header every ZIP container — and therefore every
#: XLSX — starts with. It is written from integers on purpose: an earlier
#: revision spelled it ``b"PK\\x03\\x04"``, a ten-byte string of literal
#: backslashes that no real workbook can ever begin with, and the adapter
#: silently rejected every valid export. ``test_official_records.py`` asserts
#: this constant is exactly those four bytes so the mistake cannot return.
ZIP_LOCAL_FILE_HEADER = bytes((0x50, 0x4B, 0x03, 0x04))
#: The macro part in a macro-enabled OOXML package, wherever it sits. openpyxl
#: only populates ``vba_archive`` when the workbook was loaded with ``keep_vba``,
#: so the package is inspected directly instead of trusting that attribute.
MACRO_PART = "vbaproject.bin"

HEADERS = (
    "Sr.No", "Recall Id", "FBO Name", "Brand Name", "Batch / Lot No.", "Product",
    "Reason for Recall", "Recall Start Date", "Recall Status", "Recall Termination Date",
    "License / Registration No.", "License Type [Central/State/Registration]", "Nature of Recall",
)

# Closed, reviewable failure vocabulary. Nothing else may reach
# ``OfficialSourceFetch.error_code``: openpyxl, zipfile and XML parser messages
# carry arbitrary text and are not an audit record.
ERROR_UNSUPPORTED_EXPORT = "unsupported_official_export"
ERROR_INVALID_EXPORT = "invalid_official_export"
ERROR_UNEXPECTED_SCHEMA = "unexpected_official_export_schema"
ERROR_MACRO_EXPORT = "macro_enabled_official_export"
ERROR_EMPTY_EXPORT = "empty_official_export"
ERROR_MISSING_RECORD_ID = "missing_official_record_id"
ERROR_DUPLICATE_RECORD_ID = "duplicate_official_record_id"
ERROR_MALFORMED_IDENTIFIER = "malformed_official_identifier"
ERROR_MALFORMED_DATE = "malformed_official_date"
ERROR_OUT_OF_ORDER_SOURCE_CHECK = "out_of_order_source_check"
ERROR_DUPLICATE_SOURCE_CHECK = "duplicate_source_check"
ERROR_CONFLICTING_SOURCE_CHECK = "conflicting_source_check"
ERROR_UNHANDLED = "unhandled_official_source_error"

SOURCE_ERROR_CODES = frozenset({
    ERROR_UNSUPPORTED_EXPORT, ERROR_INVALID_EXPORT, ERROR_UNEXPECTED_SCHEMA, ERROR_MACRO_EXPORT,
    ERROR_EMPTY_EXPORT, ERROR_MISSING_RECORD_ID, ERROR_DUPLICATE_RECORD_ID, ERROR_MALFORMED_IDENTIFIER,
    ERROR_MALFORMED_DATE, ERROR_OUT_OF_ORDER_SOURCE_CHECK, ERROR_DUPLICATE_SOURCE_CHECK,
    ERROR_CONFLICTING_SOURCE_CHECK, ERROR_UNHANDLED,
})

#: Values a FoSCoS export uses to say "this pack carries no lot identifier".
#: They are printed in the batch column but identify nothing, so matching a
#: pack on one would attach a stranger's recall to a customer's shelf.
BATCH_PLACEHOLDERS = frozenset({
    "na", "n/a", "nil", "none", "not applicable", "not available", "other", "others",
    "-", ".", "no", "loose", "loose sample", "sold as loose",
})


class SourceError(ValueError):
    """A controlled Step 4 source failure, carrying whatever provenance survived.

    ``source_file_sha256`` is present whenever the original bytes were readable,
    so a rejected official artifact stays auditable in the failure ledger.
    """

    def __init__(self, code: str, *, source_file_sha256: str | None = None) -> None:
        super().__init__(code)
        self.code = code if code in SOURCE_ERROR_CODES else ERROR_UNHANDLED
        self.source_file_sha256 = source_file_sha256


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(unicodedata.normalize("NFKC", str(value)).split())
    return text or None


def _identifier(value: Any, digest: str | None) -> str | None:
    """Read one official identifier cell without letting a type change rewrite it.

    Licence and batch/lot must be genuine text. A numeric cell reaching
    ``str(value)`` turns "0789" into "789" and a licence into
    "10012345678901.0"; either damages an official identifier, so the import is
    refused rather than the identity quietly rewritten.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if not isinstance(value, str):
        raise SourceError(ERROR_MALFORMED_IDENTIFIER, source_file_sha256=digest)
    return _text(value)


def _record_id(value: Any, digest: str | None) -> str:
    """The observed source stores Recall Id as a number; keep it exactly as printed.

    An integral numeric cell becomes its integer form. A fractional one is not a
    record identifier, and neither is a blank.
    """
    if isinstance(value, bool):
        raise SourceError(ERROR_MALFORMED_IDENTIFIER, source_file_sha256=digest)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not value.is_integer():
            raise SourceError(ERROR_MALFORMED_IDENTIFIER, source_file_sha256=digest)
        return str(int(value))
    if value is not None and not isinstance(value, str):
        raise SourceError(ERROR_MALFORMED_IDENTIFIER, source_file_sha256=digest)
    text = _text(value)
    if text is None:
        raise SourceError(ERROR_MISSING_RECORD_ID, source_file_sha256=digest)
    return text


def _date(value: Any, *, optional: bool = False, digest: str | None = None) -> date | None:
    """DD-MM-YYYY text, as the public export prints it. A typed date cell is honoured."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if text is None or (optional and text.casefold() == "na"):
        return None
    try:
        return datetime.strptime(text, "%d-%m-%Y").date()
    except ValueError as exc:
        raise SourceError(ERROR_MALFORMED_DATE, source_file_sha256=digest) from exc


def canonical_row(row: dict[str, Any], *, digest: str | None = None) -> dict[str, Any]:
    return {
        "external_record_id": _record_id(row.get("Recall Id"), digest),
        "fbo_name": _text(row.get("FBO Name")),
        "brand_name": _text(row.get("Brand Name")),
        "batch_lot": _identifier(row.get("Batch / Lot No."), digest),
        "product_name": _text(row.get("Product")),
        "reason": _text(row.get("Reason for Recall")),
        "recall_start_date": _date(row.get("Recall Start Date"), digest=digest),
        "recall_status": _text(row.get("Recall Status")),
        "recall_termination_date": _date(row.get("Recall Termination Date"), optional=True, digest=digest),
        "licence": _identifier(row.get("License / Registration No."), digest),
        "license_type": _text(row.get("License Type [Central/State/Registration]")),
        "nature_of_recall": _text(row.get("Nature of Recall")),
    }


def _reject_macro_package(path: Path, digest: str) -> None:
    """A macro-enabled package renamed to .xlsx is still a macro-enabled package."""
    try:
        with zipfile.ZipFile(path) as package:
            names = {name.casefold() for name in package.namelist()}
    except (zipfile.BadZipFile, OSError, ValueError) as exc:
        raise SourceError(ERROR_UNSUPPORTED_EXPORT, source_file_sha256=digest) from exc
    if any(name == MACRO_PART or name.endswith(f"/{MACRO_PART}") for name in names):
        raise SourceError(ERROR_MACRO_EXPORT, source_file_sha256=digest)


def parse_recall_xlsx(path: Path) -> tuple[list[dict[str, Any]], str]:
    """Parse the public FoSCoS Export to excel artifact without executing content.

    Every failure raises :class:`SourceError` with a closed code, so the ledger
    never records a parser's free text.
    """
    if path.suffix.casefold() != ".xlsx" or not path.is_file() or path.stat().st_size > MAX_SOURCE_BYTES:
        raise SourceError(ERROR_UNSUPPORTED_EXPORT)
    source_bytes = path.read_bytes()
    digest = hashlib.sha256(source_bytes).hexdigest()
    # Both halves matter: the signature rejects a CSV renamed .xlsx immediately,
    # and is_zipfile rejects a container whose directory is truncated or corrupt.
    if not source_bytes.startswith(ZIP_LOCAL_FILE_HEADER) or not zipfile.is_zipfile(path):
        raise SourceError(ERROR_UNSUPPORTED_EXPORT, source_file_sha256=digest)
    _reject_macro_package(path, digest)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise SourceError(ERROR_UNSUPPORTED_EXPORT, source_file_sha256=digest) from exc
    try:
        if workbook.sheetnames != [SHEET_NAME] or workbook.vba_archive is not None:
            raise SourceError(ERROR_UNSUPPORTED_EXPORT, source_file_sha256=digest)
        rows = list(workbook[SHEET_NAME].iter_rows(values_only=False))
    finally:
        workbook.close()
    if not rows or len(rows) - 1 > MAX_SOURCE_ROWS:
        raise SourceError(ERROR_INVALID_EXPORT, source_file_sha256=digest)
    if tuple(cell.value for cell in rows[0]) != HEADERS:
        raise SourceError(ERROR_UNEXPECTED_SCHEMA, source_file_sha256=digest)
    parsed: list[dict[str, Any]] = []
    seen: set[str] = set()
    for cells in rows[1:]:
        if any(cell.data_type == "f" for cell in cells):
            raise SourceError(ERROR_INVALID_EXPORT, source_file_sha256=digest)
        if len(cells) != len(HEADERS):
            raise SourceError(ERROR_UNEXPECTED_SCHEMA, source_file_sha256=digest)
        row = {HEADERS[index]: cell.value for index, cell in enumerate(cells)}
        if all(value is None for value in row.values()):
            continue
        canonical = canonical_row(row, digest=digest)
        # One workbook may not carry the same Recall Id twice: last-row-wins would
        # pick a winner the source never named, and two revisions from one source
        # check would invent a history. The whole artifact is refused instead.
        if canonical["external_record_id"] in seen:
            raise SourceError(ERROR_DUPLICATE_RECORD_ID, source_file_sha256=digest)
        seen.add(canonical["external_record_id"])
        parsed.append(canonical)
    # Headers with no rows is a successful-looking export that says nothing. Were
    # it accepted it would advance last_successful_check_at and make a stale copy
    # of the official register look current.
    if not parsed:
        raise SourceError(ERROR_EMPTY_EXPORT, source_file_sha256=digest)
    return parsed, digest


def stable_content_hash(payload: dict[str, Any]) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(encoded.encode()).hexdigest()


def normalise_licence(value: Any) -> str | None:
    """The fourteen printed digits of an FSSAI licence, or nothing.

    Deleting every non-digit from arbitrary text would collapse
    ``FSSAI 10012345678901``, ``10012345678901X`` and ``10012-345678901`` onto
    one exact official identifier and invent an exactness the source never
    stated. Only the digits themselves survive — optionally grouped with spaces,
    because labels print them that way and the confirmed-label extraction keeps
    what it read. Anything else is not a licence we can match on.
    """
    text = _text(value)
    if text is None:
        return None
    compact = text.replace(" ", "")
    if len(compact) != LICENCE_LENGTH or not compact.isdigit():
        return None
    # A run of one repeated digit is a placeholder, not a licence.
    return compact if is_valid_licence(compact) else None


def normalise_batch(value: Any) -> str | None:
    """A batch/lot identifier, or None when the source printed a placeholder.

    Whitespace is collapsed and case folded; separators are left alone, so
    ``B-123`` and ``B 123`` stay different identifiers. Short real lots such as
    ``C`` or ``1`` remain matchable — only the closed placeholder vocabulary and
    zero-only strings are refused.
    """
    text = _text(value)
    if not text:
        return None
    normalized = text.casefold()
    if normalized in BATCH_PLACEHOLDERS:
        return None
    return None if set(normalized) == {"0"} else normalized


def normalise_identity_text(value: Any) -> str | None:
    """Brand and product text for conflict detection only.

    Deliberately separate from :func:`normalise_batch`: a brand legitimately
    named ``Other`` or a product called ``0`` is a name, not a missing lot
    number, and the batch placeholder vocabulary has no meaning here.
    """
    text = _text(value)
    return text.casefold() if text else None
